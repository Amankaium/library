from flask import Flask, render_template, request
from openpyxl import load_workbook
from sqlalchemy.orm.session import sessionmaker
from sqlalchemy.orm import scoped_session
from database import engine, Book

app = Flask(__name__)
db = scoped_session(sessionmaker(bind=engine))

@app.route("/")
def homepage():
    return render_template("index.html")


@app.route("/form/")
def form():
    return render_template("form.html")



@app.route("/books/")
def books():
    # v1
    # tales = [tale.value for tale in page["A"]][1:]
    # tales = []
    # for tale in page["A"][1:]:
    #     tales.append(tale.value)

    # authors = [author.value for author in page["B"]][1:]   

    # html = """
    #     <a href="/authors">Авторы</a>
    #     <a href="/books">Книги</a>
    #     <h1 style="color: red">Тут будет список книг:</h1>
    # """

    # for i in range(len(tales)):
    #     html += f"<h2>{tales[i]} - {authors[i]}</h2>"

    # return html

    # v_2
    # excel = load_workbook("tales.xlsx")
    # page = excel["Лист1"]

    # object_list = [[tale.value, tale.offset(column=1).value] for tale in page["A"][1:]]
    # return render_template("books.html", object_list=object_list)

    # v_3
    # session = sessionmaker(engine)()
    # books = session.query(Book)
    # session.commit()

    with engine.connect() as con:
        books = con.execute("""SELECT * FROM "Book";""")

    return render_template("books_table.html", object_list=books)


@app.route("/authors/")
def authors():
    excel = load_workbook("tales.xlsx")
    page = excel["Лист1"]
    authors = {author.value for author in page["B"][1:]}
    return render_template(
        "authors.html", authors=list(authors)
    )

@app.route("/db/authors/")
def db_authors():
    with engine.connect() as con:
        authors = con.execute('SELECT DISTINCT author FROM "Book";')
 
    return render_template(
        "database_authors.html", authors=authors
    )


@app.route("/add/", methods=["POST"])
def add():
    f = request.form
    # print(f["author"], f["book"])
    excel = load_workbook("tales.xlsx")
    page = excel["Лист1"]
    last = len(page["A"]) + 1
    page[f"A{last}"] = f["book"]
    page[f"B{last}"] = f["author"]
    excel.save("tales.xlsx")
    return "форма получена"


@app.route("/db/add/", methods=["POST"])
def db_add():
    f = request.form
    book = f["book"]
    author = f["author"]
    url = f["url"]    
    
    ids = db.execute('SELECT id FROM "Book" ORDER BY id DESC;')
    max_id = ids.first().id
    c_id = max_id + 1

    db.execute(f'''
        INSERT INTO "Book" (id, name, author, image)
        VALUES ({c_id}, '{book}', '{author}', '{url}');''')
    
    db.commit()

    return "форма получена"


@app.route("/book/<num>/") # 5
def book(num): # 5
    excel = load_workbook("tales.xlsx")
    page = excel["Лист1"]
    object_list = [[tale.value, tale.offset(column=1).value, tale.offset(column=2).value] for tale in page["A"][1:]]
    obj = object_list[int(num)] # object_list[5]
    obj.append(num)
    return render_template("book.html", obj=obj) # **kwargs


@app.route("/book/<num>/edit/")
def book_edit(num):
    num = int(num) + 2
    excel_file = load_workbook("tales.xlsx")
    page = excel_file["Лист1"]
    tale = page[f"A{num}"]
    author = page[f"B{num}"]
    image = page[f"C{num}"]
    obj = [tale.value, author.value, image.value, num]
    return render_template("book_edit.html", obj=obj)


@app.route("/book/<num>/save/", methods=["POST"])
def book_save(num):
    num = int(num)
    excel_file = load_workbook("tales.xlsx")
    page = excel_file["Лист1"]
    form = request.form
    page[f"A{num}"] = form["tale"]
    page[f"B{num}"] = form["author"]
    page[f"C{num}"] = form["image"]
    excel_file.save("tales.xlsx")
    return "Сохранено!"